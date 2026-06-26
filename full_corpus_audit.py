"""
Full corpus audit — every file in D:/Test_files tested against the live pipeline.

Optimised for low-end hardware (1 GHz single-core, 2 GB RAM) and scales up
automatically on better machines:

  Workers      auto-detect: floor(logical_cpus * 0.70) capped by free RAM / 700 MB
  Sort order   smallest files first; large files last
  Large ZIPs   extracted to temp dir; each member tested individually (no OOM)
  Large PDFs   split into 20-page chunks; each chunk tested separately
  Large XLSX   split into per-sheet temp files; each sheet tested separately
  RAM throttle dispatching pauses if system RAM exceeds --ram-cap (default 88%)
  Worker       single ONNX session per lifetime (not reloaded per file)
               gc.collect() after every file; thread counts capped to CPU budget

GPU note: Intel UHD integrated graphics detected. DirectML ONNX provider is NOT
available in this build (onnxruntime 1.27.0, CPUExecutionProvider only). Text
extraction is CPU-bound (XML/string parsing) and does not benefit from GPU.
GPU acceleration would only apply to OCR — which is disabled with --no-ocr.

Usage:
    python full_corpus_audit.py [--no-ocr] [--workers N|auto] [--out FILE]
                                [--file-timeout N] [--ram-cap N]
                                [--expand-threshold MB] [--pdf-chunk-pages N]
"""
import sys, io, json, time, pathlib, argparse, subprocess, threading, queue, tempfile, zipfile
from collections import Counter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

TEST_ROOT = pathlib.Path("D:/Test_files")
SKIP_NAME_PREFIXES = ("~$", "._", ".~")
SKIP_EXTS = {".tmp", ".ds_store", ".lnk", ".bak", ""}

# ---------------------------------------------------------------------------
# Persistent worker: reads JSON from stdin, writes JSON to stdout.
# One subprocess per slot — ONNX/magika model loaded ONCE, not once per file.
# ---------------------------------------------------------------------------
_WORKER_CODE = """\
import sys, json, time, pathlib, gc, os
# ── Thread budget for single-core / memory-constrained machines ─────────────
_threads = sys.argv[2] if len(sys.argv) > 2 else "1"
os.environ.setdefault("OMP_NUM_THREADS",          _threads)
os.environ.setdefault("OPENBLAS_NUM_THREADS",     _threads)
os.environ.setdefault("MKL_NUM_THREADS",          _threads)
os.environ.setdefault("NUMEXPR_NUM_THREADS",      _threads)
os.environ.setdefault("VECLIB_MAXIMUM_THREADS",   _threads)
os.environ.setdefault("TF_NUM_INTRAOP_THREADS",   _threads)
os.environ.setdefault("TF_NUM_INTEROP_THREADS",   _threads)
# ── ONNX GPU providers: use DirectML/CUDA if available ─────────────────────
try:
    import onnxruntime as _ort
    _pv = _ort.get_available_providers()
    for _p in ("DmlExecutionProvider", "CUDAExecutionProvider", "ROCMExecutionProvider"):
        if _p in _pv:
            os.environ.setdefault("ORT_DEFAULT_PROVIDER", _p)
            break
except ImportError:
    pass
# ── Pipeline import + ONNX pre-warm (magika model loaded before WORKER_READY) ─
proj = sys.argv[1]
sys.path.insert(0, proj)
from pipeline import convert_file, _get_markitdown
try:
    _get_markitdown()   # loads MarkItDown + magika ONNX (~8-15 s, done once here)
except Exception:
    pass
sys.stderr.write("WORKER_READY\\n")
sys.stderr.flush()
# ── Main loop ───────────────────────────────────────────────────────────────
for raw in sys.stdin:
    raw = raw.strip()
    if not raw or raw == "STOP":
        break
    t0 = time.time()
    ext = ""
    kb = 0.0
    path_str = "?"
    try:
        req = json.loads(raw)
        path_str = req["path"]
        auto_ocr = req.get("auto_ocr", False)
        p = pathlib.Path(path_str)
        ext = p.suffix.lower()
        kb = round(p.stat().st_size / 1024, 1) if p.exists() else 0.0
        r = convert_file(path_str, auto_ocr=auto_ocr, auto_bijoy=True)
        txt = r.get("text", "")
        words = len(txt.split()) if txt else 0
        status = "PASS" if words > 0 else "EMPTY"
        result = {"status": status, "ext": ext, "kb": kb,
                  "elapsed": round(time.time() - t0, 2), "words": words,
                  "steps": r.get("steps", []), "file": path_str,
                  "name": p.name, "error": ""}
    except MemoryError as exc:
        result = {"status": "FAIL", "ext": ext, "kb": kb,
                  "elapsed": round(time.time() - t0, 2), "words": 0,
                  "steps": [], "file": path_str,
                  "name": pathlib.Path(path_str).name,
                  "error": "MemoryError: " + str(exc)[:200]}
    except Exception as exc:
        msg = str(exc)
        s = ("UNSUPPORTED" if "unsupported format" in msg.lower() or
             "no text can be extracted" in msg.lower() else "FAIL")
        result = {"status": s, "ext": ext, "kb": kb,
                  "elapsed": round(time.time() - t0, 2), "words": 0,
                  "steps": [], "file": path_str,
                  "name": pathlib.Path(path_str).name, "error": msg[:300]}
    print(json.dumps(result), flush=True)
    gc.collect()
"""


# ---------------------------------------------------------------------------
# Resource detection
# ---------------------------------------------------------------------------

def _auto_workers() -> int:
    """Safe worker count: 70% of logical CPUs, capped by free RAM (700 MB/worker)."""
    try:
        import psutil
        cpu_limit = max(1, int(psutil.cpu_count(logical=True) * 0.70))
        free_gb = psutil.virtual_memory().available / (1024 ** 3)
        ram_limit = max(1, int(free_gb / 0.70))   # 700 MB per worker slot
        return min(cpu_limit, ram_limit)
    except ImportError:
        return 1


def _cpu_threads_per_worker(n_workers: int) -> int:
    """Per-worker thread budget = floor(70% of logical CPUs / workers), min 1."""
    try:
        import psutil
        total = psutil.cpu_count(logical=True) or 1
        return max(1, int(total * 0.70 / n_workers))
    except ImportError:
        return 1


def _detect_gpu() -> str:
    try:
        import onnxruntime as ort
        pv = ort.get_available_providers()
        for p in ("DmlExecutionProvider", "CUDAExecutionProvider", "ROCMExecutionProvider"):
            if p in pv:
                return f"GPU — ONNX provider: {p} (onnxruntime {ort.__version__})"
        return f"CPU only (no DirectML/CUDA in onnxruntime {ort.__version__})"
    except ImportError:
        return "onnxruntime not importable"


# ---------------------------------------------------------------------------
# Large-file expansion (ZIP extraction + PDF page-splitting)
# ---------------------------------------------------------------------------

def _extract_zip_members(zip_path: str, tmp_dir: pathlib.Path) -> list:
    """Extract a ZIP to a temp sub-dir; return list of extracted file paths."""
    p = pathlib.Path(zip_path)
    out_dir = tmp_dir / f"zip_{p.stem[:40]}"
    out_dir.mkdir(parents=True, exist_ok=True)
    extracted = []
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            for member in zf.infolist():
                if member.is_dir():
                    continue
                name = pathlib.Path(member.filename).name
                ext  = pathlib.Path(member.filename).suffix.lower()
                if any(name.startswith(pf) for pf in SKIP_NAME_PREFIXES):
                    continue
                if ext in SKIP_EXTS or name.lower() == ".ds_store":
                    continue
                out = zf.extract(member, out_dir)
                if pathlib.Path(out).exists():   # silent failure on long Windows paths
                    extracted.append(str(out))
    except Exception:
        pass
    return extracted


def _split_pdf(pdf_path: str, tmp_dir: pathlib.Path, chunk_pages: int) -> list:
    """Split a PDF into chunk_pages-page temp files; return list of paths."""
    try:
        import fitz
    except ImportError:
        return []
    p = pathlib.Path(pdf_path)
    try:
        doc = fitz.open(pdf_path)
        total = len(doc)
        if total <= chunk_pages:
            doc.close()
            return []           # small enough; don't split
        out_dir = tmp_dir / f"pdf_{p.stem[:40]}"
        out_dir.mkdir(parents=True, exist_ok=True)
        chunks = []
        for i, start in enumerate(range(0, total, chunk_pages)):
            end = min(start + chunk_pages, total)
            out_p = out_dir / f"chunk_{i+1:04d}_p{start+1}-{end}.pdf"
            chunk = fitz.open()
            chunk.insert_pdf(doc, from_page=start, to_page=end - 1)
            chunk.save(str(out_p))
            chunk.close()
            chunks.append(str(out_p))
        doc.close()
        return chunks
    except Exception:
        return []


def _split_xlsx(xlsx_path: str, tmp_dir: pathlib.Path) -> list:
    """Split an XLSX into one temp file per sheet. Returns list of paths."""
    try:
        import openpyxl, re as _re
    except ImportError:
        return []
    try:
        p = pathlib.Path(xlsx_path)
        # peek at sheet names without loading data
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        if len(sheet_names) <= 1:
            return []   # single-sheet; no point splitting
        out_dir = tmp_dir / f"xlsx_{p.stem[:40]}"
        out_dir.mkdir(parents=True, exist_ok=True)
        chunks = []
        for i, name in enumerate(sheet_names):
            wb_in  = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws_in  = wb_in[name]
            wb_out = openpyxl.Workbook(write_only=True)
            ws_out = wb_out.create_sheet(name)
            for row in ws_in.iter_rows(values_only=True):
                ws_out.append([v for v in row])
            safe  = _re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name)[:30]
            out_p = out_dir / f"sheet_{i+1:03d}_{safe}.xlsx"
            wb_out.save(str(out_p))
            wb_in.close()
            if pathlib.Path(out_p).exists():
                chunks.append(str(out_p))
        return chunks
    except Exception:
        return []


def _expand_files(
    all_files: list,
    tmp_dir: pathlib.Path,
    threshold_mb: float,
    pdf_chunk_pages: int,
) -> tuple:
    """
    Replace large ZIPs / PDFs / XLSX with their smaller constituents.
    Returns (expanded_list, split_log).
    Files that cannot be expanded are kept as-is (sorted last via size sort).
    """
    split_log = []
    result_small = []   # files that were expanded or are small
    result_large = []   # large files that couldn't be expanded (process last)
    threshold_bytes = threshold_mb * 1024 * 1024

    for path in all_files:
        p = pathlib.Path(path)
        try:
            size = p.stat().st_size
        except OSError:
            result_small.append(path)
            continue
        ext = p.suffix.lower()

        if ext == ".zip" and size > threshold_bytes:
            members = _extract_zip_members(path, tmp_dir)
            if members:
                split_log.append(
                    f"  ZIP  {p.name} ({size/1024**2:.1f} MB) → {len(members)} members"
                )
                result_small.extend(members)
            else:
                split_log.append(f"  ZIP  {p.name} ({size/1024**2:.1f} MB) → extract failed, kept")
                result_large.append(path)
        elif ext == ".pdf" and size > threshold_bytes:
            chunks = _split_pdf(path, tmp_dir, pdf_chunk_pages)
            if chunks:
                split_log.append(
                    f"  PDF  {p.name} ({size/1024**2:.1f} MB) → {len(chunks)} chunks of {pdf_chunk_pages} pages"
                )
                result_small.extend(chunks)
            else:
                split_log.append(f"  PDF  {p.name} ({size/1024**2:.1f} MB) → split failed, kept whole")
                result_large.append(path)
        elif ext == ".xlsx" and size > threshold_bytes:
            sheets = _split_xlsx(path, tmp_dir)
            if sheets:
                split_log.append(
                    f"  XLSX {p.name} ({size/1024**2:.1f} MB) → {len(sheets)} sheet files"
                )
                result_small.extend(sheets)
            else:
                split_log.append(f"  XLSX {p.name} ({size/1024**2:.1f} MB) → split failed, kept whole")
                result_large.append(path)
        else:
            if size > threshold_bytes:
                result_large.append(path)   # large unsplittable → end of queue
            else:
                result_small.append(path)

    # sort small/extracted by size ascending; large by size ascending too
    def _key(p_str):
        try:
            return pathlib.Path(p_str).stat().st_size
        except OSError:
            return 0

    result_small.sort(key=_key)
    result_large.sort(key=_key)
    return result_small + result_large, split_log


# ---------------------------------------------------------------------------
# Timeout helper
# ---------------------------------------------------------------------------

def _file_timeout(size_bytes: int, base: int) -> int:
    """base + 1 extra second per MB beyond the first megabyte."""
    extra = max(0, int(size_bytes / (1024 * 1024)) - 1)
    return base + extra


# ---------------------------------------------------------------------------
# Persistent worker subprocess
# ---------------------------------------------------------------------------

class WorkerSubprocess:
    STARTUP_TIMEOUT = 90   # covers ONNX pre-warm (8-15 s) + startup

    def __init__(self, proj_path: str, cpu_threads: int = 1):
        self.alive = False
        self._proc = subprocess.Popen(
            [sys.executable, "-c", _WORKER_CODE, proj_path, str(cpu_threads)],
            stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
            text=True, encoding="utf-8", errors="replace",
        )
        self._ready_event = threading.Event()
        self._stderr_thread = threading.Thread(target=self._drain_stderr, daemon=True)
        self._stderr_thread.start()
        if self._ready_event.wait(timeout=self.STARTUP_TIMEOUT):
            self.alive = True
        else:
            try:
                self._proc.kill()
            except Exception:
                pass

    def _drain_stderr(self):
        try:
            for line in self._proc.stderr:
                if "WORKER_READY" in line:
                    self._ready_event.set()
        except Exception:
            pass

    def kill(self):
        self.alive = False
        try:
            self._proc.kill()
        except Exception:
            pass

    def process(self, path: str, auto_ocr: bool, timeout: int):
        """Returns (result_dict, None) or (None, error_str)."""
        resp_q: queue.Queue = queue.Queue()

        def _read():
            try:
                line = self._proc.stdout.readline()
                resp_q.put(line.strip() if line else "")
            except Exception:
                resp_q.put(None)

        try:
            self._proc.stdin.write(json.dumps({"path": path, "auto_ocr": auto_ocr}) + "\n")
            self._proc.stdin.flush()
        except Exception as exc:
            self.alive = False
            return None, f"stdin write: {exc}"

        reader = threading.Thread(target=_read, daemon=True)
        reader.start()
        reader.join(timeout=timeout)

        if reader.is_alive():
            self.kill()
            return None, f"Timeout after {timeout}s"

        try:
            data = resp_q.get_nowait()
        except queue.Empty:
            self.alive = False
            return None, "No response"

        if data is None or data == "":
            self.alive = False
            return None, "Worker stdout closed (crashed)"

        try:
            return json.loads(data), None
        except json.JSONDecodeError:
            self.alive = False
            return None, f"Bad JSON: {data[:80]}"


# ---------------------------------------------------------------------------
# Progress + RAM throttle
# ---------------------------------------------------------------------------

_print_lock  = threading.Lock()
_last_report = [0.0]


def _progress(done: int, total: int, t_start: float) -> None:
    now = time.time()
    if done == total or done % 50 == 0 or (now - _last_report[0]) >= 30:
        with _print_lock:
            now2 = time.time()
            if done == total or done % 50 == 0 or (now2 - _last_report[0]) >= 30:
                elapsed = round(now2 - t_start, 1)
                rate = done / elapsed if elapsed > 0 else 0
                eta = round((total - done) / rate) if rate > 0 else 0
                print(f"  [{done}/{total}] {elapsed}s elapsed  rate={rate:.2f}/s  "
                      f"ETA={eta}s …", flush=True)
                _last_report[0] = now2


def _throttle_ram(ram_cap_pct: int, interval: float = 1.5, max_wait: float = 120) -> None:
    """Sleep until system RAM usage drops below cap (or max_wait seconds pass)."""
    try:
        import psutil
    except ImportError:
        return
    waited = 0.0
    while psutil.virtual_memory().percent >= ram_cap_pct and waited < max_wait:
        time.sleep(interval)
        waited += interval


# ---------------------------------------------------------------------------
# Worker slot thread
# ---------------------------------------------------------------------------

def _worker_slot(
    file_q: queue.Queue,
    results: list,
    results_lock: threading.Lock,
    proj: str,
    base_timeout: int,
    ram_cap: int,
    counter: list,
    total: int,
    t_start: float,
    cpu_threads: int,
) -> None:
    MAX_RESTARTS = 5
    restarts = 0
    worker = WorkerSubprocess(proj, cpu_threads)

    while True:
        item = file_q.get()
        if item is None:
            file_q.task_done()
            break

        path, auto_ocr = item
        p = pathlib.Path(path)
        try:
            size = p.stat().st_size
        except OSError:
            size = 0
        ext = p.suffix.lower()
        kb  = round(size / 1024, 1)

        # RAM throttle — pause before dispatching if system is memory-pressured
        _throttle_ram(ram_cap)

        timeout = _file_timeout(size, base_timeout)

        # Restart dead worker
        if not worker.alive:
            if restarts >= MAX_RESTARTS:
                r = {"status": "FAIL", "ext": ext, "kb": kb, "elapsed": 0.0,
                     "steps": [], "words": 0, "file": path, "name": p.name,
                     "error": f"Worker dead after {MAX_RESTARTS} restarts"}
                with results_lock:
                    results.append(r); counter[0] += 1; done = counter[0]
                _progress(done, total, t_start)
                file_q.task_done()
                continue
            restarts += 1
            with _print_lock:
                print(f"  [restart {restarts}/{MAX_RESTARTS}] spawning new worker …",
                      flush=True)
            worker = WorkerSubprocess(proj, cpu_threads)
            if not worker.alive:
                r = {"status": "FAIL", "ext": ext, "kb": kb, "elapsed": 0.0,
                     "steps": [], "words": 0, "file": path, "name": p.name,
                     "error": "New worker failed to start (60 s timeout)"}
                with results_lock:
                    results.append(r); counter[0] += 1; done = counter[0]
                _progress(done, total, t_start)
                file_q.task_done()
                continue

        result, err = worker.process(path, auto_ocr, timeout)

        if result is not None:
            r = result
            restarts = 0
        else:
            elapsed_val = float(timeout) if "Timeout" in (err or "") else 0.0
            r = {"status": "FAIL", "ext": ext, "kb": kb, "elapsed": elapsed_val,
                 "steps": [], "words": 0, "file": path, "name": p.name,
                 "error": err or "Unknown error"}

        with results_lock:
            results.append(r); counter[0] += 1; done = counter[0]
        _progress(done, total, t_start)
        file_q.task_done()

    worker.kill()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Full corpus audit against the live pipeline.")
    ap.add_argument("--no-ocr", action="store_true",
                    help="Disable OCR (faster; tests text extraction only)")
    ap.add_argument("--workers", default="auto",
                    help="Worker count or 'auto' (default auto — 70%% CPU, capped by free RAM)")
    ap.add_argument("--out", default="full_audit_results.json")
    ap.add_argument("--file-timeout", type=int, default=120,
                    help="Base per-file timeout in seconds; scales +1s/MB (default 120)")
    ap.add_argument("--ram-cap", type=int, default=88,
                    help="Pause dispatch if system RAM %% >= this value (default 88)")
    ap.add_argument("--expand-threshold", type=float, default=20.0,
                    help="Expand ZIPs / split PDFs/XLSX above this size in MB (default 20)")
    ap.add_argument("--pdf-chunk-pages", type=int, default=20,
                    help="Pages per PDF chunk when splitting (default 50)")
    args = ap.parse_args()

    auto_ocr = not args.no_ocr
    proj_dir = str(pathlib.Path(__file__).parent)

    # ── Worker count ─────────────────────────────────────────────────────────
    if args.workers == "auto":
        n_workers = _auto_workers()
    else:
        n_workers = max(1, int(args.workers))

    cpu_threads = _cpu_threads_per_worker(n_workers)
    gpu_info    = _detect_gpu()

    # ── Collect all files ────────────────────────────────────────────────────
    raw_files = []
    for f in TEST_ROOT.rglob("*"):
        if not f.is_file():
            continue
        if any(f.name.startswith(pf) for pf in SKIP_NAME_PREFIXES):
            continue
        if f.suffix.lower() in SKIP_EXTS or f.name.lower() == ".ds_store":
            continue
        raw_files.append(str(f))

    # ── Expand large ZIPs / split large PDFs ─────────────────────────────────
    tmp_dir = pathlib.Path(tempfile.mkdtemp(prefix="audit_expand_"))
    try:
        all_files, split_log = _expand_files(
            raw_files, tmp_dir, args.expand_threshold, args.pdf_chunk_pages
        )

        total = len(all_files)
        print(f"\nFull corpus audit — {len(raw_files)} source files → {total} test items")
        print(f"  Workers: {n_workers} (auto)  •  CPU threads/worker: {cpu_threads}")
        print(f"  GPU: {gpu_info}")
        print(f"  RAM cap: {args.ram_cap}%  •  base timeout: {args.file_timeout}s")
        print(f"  Expand threshold: {args.expand_threshold} MB  "
              f"(PDF chunks: {args.pdf_chunk_pages} pages, XLSX by sheet)")
        print(f"  OCR: {'on' if auto_ocr else 'off'}")
        if split_log:
            print(f"  Expanded {len(split_log)} large file(s):")
            for line in split_log:
                print(line)
        print("=" * 80)

        # ── Queue ────────────────────────────────────────────────────────────
        file_q: queue.Queue = queue.Queue()
        for f in all_files:
            file_q.put((f, auto_ocr))
        for _ in range(n_workers):
            file_q.put(None)

        results: list = []
        results_lock = threading.Lock()
        counter = [0]
        t_start = time.time()

        threads = [
            threading.Thread(
                target=_worker_slot,
                args=(file_q, results, results_lock, proj_dir,
                      args.file_timeout, args.ram_cap, counter, total, t_start,
                      cpu_threads),
                daemon=True,
            )
            for _ in range(n_workers)
        ]
        for t in threads:
            t.start()
        for t in threads:
            t.join()

        # ── Write results ─────────────────────────────────────────────────────
        out_path = pathlib.Path(args.out)
        out_path.write_text(json.dumps(results, indent=2, ensure_ascii=False),
                            encoding="utf-8")

        by_status: dict = {}
        for r in results:
            by_status.setdefault(r["status"], []).append(r)

        print(f"\n{'Status':<16} {'Count':>6}  {'Ext distribution (top 8)'}")
        print("-" * 80)
        for st in ("PASS", "EMPTY", "UNSUPPORTED", "FAIL"):
            group = by_status.get(st, [])
            if not group:
                continue
            top = Counter(r["ext"] for r in group).most_common(8)
            ext_str = "  ".join(f"{e}×{c}" for e, c in top)
            print(f"{st:<16} {len(group):>6}  {ext_str}")

        fails = by_status.get("FAIL", [])
        if fails:
            print(f"\nFAIL detail ({len(fails)} files):")
            err_groups: dict = {}
            for r in fails:
                key = r["error"][:80]
                err_groups.setdefault(key, []).append(r["ext"])
            for msg, exts in sorted(err_groups.items(), key=lambda x: -len(x[1])):
                print(f"  [{len(exts)}]  {msg[:70]}")
                print(f"        {dict(Counter(exts))}")

        empties = by_status.get("EMPTY", [])
        if empties:
            print(f"\nEMPTY ({len(empties)} — converted but yielded no text):")
            for ext, cnt in Counter(r["ext"] for r in empties).most_common():
                print(f"  {ext}: {cnt}")

        total_time = round(time.time() - t_start, 1)
        pass_c  = len(by_status.get("PASS",        []))
        empty_c = len(by_status.get("EMPTY",       []))
        unsup_c = len(by_status.get("UNSUPPORTED", []))
        fail_c  = len(by_status.get("FAIL",        []))
        fixable = pass_c + empty_c + fail_c
        print(f"\nSUMMARY: {pass_c} PASS  |  {empty_c} EMPTY  "
              f"|  {unsup_c} UNSUPPORTED  |  {fail_c} FAIL")
        print(f"Fixable pass rate: {pass_c}/{fixable} = {100*pass_c//max(fixable,1)}%")
        print(f"Total time: {total_time}s  |  Results: {out_path}")
        sys.exit(0 if fail_c == 0 else 1)

    finally:
        # Clean up expanded temp files
        import shutil
        try:
            shutil.rmtree(tmp_dir, ignore_errors=True)
        except Exception:
            pass


if __name__ == "__main__":
    main()
